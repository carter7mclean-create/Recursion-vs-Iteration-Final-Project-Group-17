from __future__ import annotations

import html
import math
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
INPUT_XLSX = Path("/Users/addisonlin/Downloads/codeProjectAverages.xlsx")
OUTPUT_SVG = ROOT / "results" / "charts" / "poster_binary_search_from_excel.svg"

RECURSIVE_COLOR = "#c2410c"
ITERATIVE_COLOR = "#0f766e"
BG_COLOR = "#f8fafc"
PANEL_FILL = "#ffffff"
PANEL_STROKE = "#dbe4ee"
GRID_COLOR = "#e2e8f0"
AXIS_COLOR = "#334155"
TEXT_DARK = "#0f172a"
TEXT_MUTED = "#475569"
TEXT_LIGHT = "#64748b"


def read_binary_search_averages(path: Path) -> tuple[list[str], list[int], list[int]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    recursive_labels: list[str] = []
    recursive_avgs: list[int] = []
    iterative_avgs: list[int] = []

    mode = None
    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if first == "recursive":
            mode = "recursive"
            continue
        if first == "iterative":
            mode = "iterative"
            continue
        if first is None or mode is None:
            continue

        label = str(first).strip()
        avg_value = row[12]
        if avg_value is None:
            continue

        if mode == "recursive":
            recursive_labels.append(label)
            recursive_avgs.append(int(avg_value))
        else:
            iterative_avgs.append(int(avg_value))

    if not recursive_labels or len(recursive_labels) != len(recursive_avgs) or len(recursive_avgs) != len(iterative_avgs):
        raise ValueError("Could not parse matching recursive and iterative averages from the workbook.")

    return recursive_labels, recursive_avgs, iterative_avgs


def badge(x: float, y: float, text: str, fill: str, text_fill: str = "#ffffff", stroke: str | None = None) -> str:
    width = max(106, int(len(text) * 7.2) + 24)
    stroke_attr = f' stroke="{stroke}" stroke-width="1"' if stroke else ""
    return (
        f'<rect x="{x}" y="{y}" width="{width}" height="28" rx="14" fill="{fill}"{stroke_attr}/>'
        f'<text x="{x + 12}" y="{y + 18}" font-size="13" font-weight="700" fill="{text_fill}">{html.escape(text)}</text>'
    )


def pairwise(points: list[tuple[float, float]]) -> list[tuple[tuple[float, float], tuple[float, float]]]:
    return list(zip(points, points[1:]))


def svg_line_chart(labels: list[str], recursive: list[int], iterative: list[int], output_path: Path) -> Path:
    width = 1500
    height = 900
    panel_x = 70
    panel_y = 88
    panel_width = width - 140
    panel_height = height - 152
    plot_margin_left = 120
    plot_margin_right = 36
    plot_margin_top = 112
    plot_margin_bottom = 102
    plot_x = panel_x + plot_margin_left
    plot_y = panel_y + plot_margin_top
    plot_width = panel_width - plot_margin_left - plot_margin_right
    plot_height = panel_height - plot_margin_top - plot_margin_bottom

    values = recursive + iterative
    y_max = max(values)
    y_min = 0
    y_padding = max(500, int(y_max * 0.12))
    y_max = y_max + y_padding

    def x_pos(index: int) -> float:
        if len(labels) == 1:
            return plot_x + plot_width / 2
        return plot_x + index * plot_width / (len(labels) - 1)

    def y_pos(value: int) -> float:
        normalized = (value - y_min) / (y_max - y_min)
        return plot_y + plot_height - normalized * plot_height

    compared_ratio = recursive[-1] / iterative[-1]

    pieces = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        f'<rect width="100%" height="100%" fill="{BG_COLOR}"/>',
        f'<rect x="{panel_x}" y="{panel_y}" width="{panel_width}" height="{panel_height}" rx="30" fill="{PANEL_FILL}" stroke="{PANEL_STROKE}" stroke-width="1.5"/>',
        f'<text x="{panel_x + 26}" y="{panel_y + 40}" font-size="34" font-weight="700" fill="{TEXT_DARK}">Binary Search</text>',
        badge(panel_x + 26, panel_y + 58, "Recursive", RECURSIVE_COLOR),
        badge(panel_x + 154, panel_y + 58, "Iterative", ITERATIVE_COLOR),
        badge(panel_x + 282, panel_y + 58, "Source: teammate Excel data", "#eff6ff", text_fill="#1d4ed8", stroke="#93c5fd"),
        f'<text x="{panel_x + panel_width - 26}" y="{panel_y + 77}" text-anchor="end" font-size="18" fill="{TEXT_MUTED}">At 100k: recursion is {compared_ratio:.2f}x slower</text>',
    ]

    tick_count = 5
    tick_values = [int(round(y_max * i / tick_count)) for i in range(tick_count + 1)]
    for tick_value in tick_values:
        y = y_pos(tick_value)
        pieces.append(f'<line x1="{plot_x}" y1="{y}" x2="{plot_x + plot_width}" y2="{y}" stroke="{GRID_COLOR}" stroke-width="1.2"/>')
        pieces.append(f'<text x="{plot_x - 16}" y="{y + 5}" text-anchor="end" font-size="16" fill="{TEXT_LIGHT}">{tick_value:,}</text>')

    for idx, label in enumerate(labels):
        x = x_pos(idx)
        pieces.append(f'<line x1="{x}" y1="{plot_y}" x2="{x}" y2="{plot_y + plot_height}" stroke="#f1f5f9" stroke-width="1"/>')
        pieces.append(f'<text x="{x}" y="{plot_y + plot_height + 34}" text-anchor="middle" font-size="16" fill="{TEXT_LIGHT}">{html.escape(label)}</text>')

    pieces.append(f'<line x1="{plot_x}" y1="{plot_y + plot_height}" x2="{plot_x + plot_width}" y2="{plot_y + plot_height}" stroke="{AXIS_COLOR}" stroke-width="2"/>')
    pieces.append(f'<line x1="{plot_x}" y1="{plot_y}" x2="{plot_x}" y2="{plot_y + plot_height}" stroke="{AXIS_COLOR}" stroke-width="2"/>')
    pieces.append(f'<text x="{plot_x + plot_width / 2}" y="{plot_y + plot_height + 68}" text-anchor="middle" font-size="18" fill="{AXIS_COLOR}">Array Size</text>')
    pieces.append(
        f'<text x="{plot_x - 88}" y="{plot_y + plot_height / 2}" text-anchor="middle" font-size="18" fill="{AXIS_COLOR}" '
        f'transform="rotate(-90 {plot_x - 88} {plot_y + plot_height / 2})">Average Runtime (ns)</text>'
    )

    recursive_points = [(x_pos(i), y_pos(value)) for i, value in enumerate(recursive)]
    iterative_points = [(x_pos(i), y_pos(value)) for i, value in enumerate(iterative)]

    for start, end in pairwise(recursive_points):
        pieces.append(
            f'<line x1="{start[0]}" y1="{start[1]}" x2="{end[0]}" y2="{end[1]}" '
            f'stroke="{RECURSIVE_COLOR}" stroke-width="5" stroke-linecap="round"/>'
        )
    for start, end in pairwise(iterative_points):
        pieces.append(
            f'<line x1="{start[0]}" y1="{start[1]}" x2="{end[0]}" y2="{end[1]}" '
            f'stroke="{ITERATIVE_COLOR}" stroke-width="5" stroke-linecap="round"/>'
        )

    for x, y in recursive_points:
        pieces.append(f'<circle cx="{x}" cy="{y}" r="7" fill="{RECURSIVE_COLOR}" stroke="#ffffff" stroke-width="2.5"/>')
    for x, y in iterative_points:
        pieces.append(f'<circle cx="{x}" cy="{y}" r="7" fill="{ITERATIVE_COLOR}" stroke="#ffffff" stroke-width="2.5"/>')

    output_path.write_text("\n".join(pieces + ["</svg>"]), encoding="utf-8")
    return output_path


def main() -> None:
    labels, recursive, iterative = read_binary_search_averages(INPUT_XLSX)
    OUTPUT_SVG.parent.mkdir(parents=True, exist_ok=True)
    output_path = svg_line_chart(labels, recursive, iterative, OUTPUT_SVG)
    print(output_path)


if __name__ == "__main__":
    main()
